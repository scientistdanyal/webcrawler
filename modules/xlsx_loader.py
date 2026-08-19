from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


@dataclass(frozen=True)
class BranchRow:
    name: str
    website: str
    sheet_email: str | None


def prefer_https(url: str) -> str:
    """Normalize website URLs to https:// (upgrade http and bare domains)."""
    cleaned = url.strip()
    if not cleaned:
        return cleaned
    if not cleaned.startswith(("http://", "https://")):
        cleaned = f"https://{cleaned}"
    elif cleaned.startswith("http://"):
        cleaned = "https://" + cleaned[len("http://") :]

    parts = urlsplit(cleaned)
    return urlunsplit(
        ("https", parts.netloc, parts.path or "", parts.query, parts.fragment)
    )


def load_branches_from_xlsx(path: str | Path) -> list[BranchRow]:
    """Load branches that have a Website value from the YMCA directory."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        return []

    header_map = {
        str(col).strip().lower(): idx
        for idx, col in enumerate(header)
        if col is not None
    }

    name_idx = header_map.get("branch/association name")
    website_idx = header_map.get("website")
    email_idx = header_map.get("email")

    if name_idx is None or website_idx is None:
        raise ValueError(
            "XLSX must include 'Branch/Association Name' and 'Website' columns"
        )

    branches: list[BranchRow] = []
    for row in rows:
        name = row[name_idx]
        website = row[website_idx]
        if not name or not website:
            continue

        website_str = str(website).strip()
        if not website_str:
            continue
        website_str = prefer_https(website_str)

        sheet_email = None
        if email_idx is not None and row[email_idx]:
            sheet_email = str(row[email_idx]).strip() or None

        branches.append(
            BranchRow(
                name=str(name).strip(),
                website=website_str,
                sheet_email=sheet_email,
            )
        )

    return branches
